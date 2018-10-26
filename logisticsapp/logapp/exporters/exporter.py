import openpyxl

from datetime import date
from openpyxl.styles import Font
from tools.company import productInformation
import os
from app.conf import base as settings
import random
import string

# Set settings to import first


class ExcelExporter():
	'''
	To export data to different exports. All instances are private except the export type below.
	'''

	#general settings

	companyHeaderHeaderFont=Font(name='Arial', size=22)
	reportTitleFont=Font(name='Arial', size=14)
	printedOnFont=Font(name='Arial', size=12)
	colHeadersFont=Font(name='Arial', size=14)

	columnLetters=['A','B','C','D','E','F','G','E','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','Y','Z']

	columnsRow=8 #this is columns row. This must be less than StartingRow below. Here, we have headers such as Name, Gender
	startingRow=9 #we start printing data records from this row.
	startingColumnIndex=1 #refers to B above

	save_to=os.path.join(settings.BASE_DIR , 'static','exports','xls')



	def export(self,current_user_name,data,calledfrom,datakind,append_to_file_name='',extra=None):
		'''
		The data is the data to be put into excel file. Calledfrom is a descriptive string that will determine which exporter to use. datakina
		is related to calledfrom and determines what kind of displayis presented. the append_to_file_name, if provided, will be part of the file name that was generated

		export: to export inventory records: data,inventory,q to mean q=>quantified.

		all functions below return a filename

		'''

		self.data=data
		self.kind=datakind
		self.append_to_file_name=append_to_file_name
		self.extra=extra
		self.calledfrom=calledfrom
		self.current_user_name=current_user_name

		


		if calledfrom=='inventory':
			file_name=self.__inventory()

		elif calledfrom=='finishing_war':
			file_name=self.__finishing_items()

		elif calledfrom=='expiring':
			file_name=self.__expiring()


		return file_name[1] #this holds URL to the file now


	def __workbook(self,title):
		#create a workbook. Called from method export only
		product_information=productInformation()
		self.wb = openpyxl.Workbook()
		self.sheet=self.wb.active
		self.sheet.title = 'Sheet 1'	
		companyHeaderCell=self.sheet['B2']
		dateCell=self.sheet['E6']
		rptTitle=self.sheet['B4']
		self.sheet.merge_cells('B2:N2') #Company name
		self.sheet.merge_cells('B4:D4') #Inventory Title
		self.sheet.merge_cells('E6:G6') #PLACE DATE HERE
		companyHeaderCell.font = self.companyHeaderHeaderFont
		dateCell.font=self.printedOnFont
		rptTitle.font=self.reportTitleFont
		rptTitle.value=title
		dateCell.value=date.today()
		companyHeaderCell.value = product_information['company_name'];

	def __buildColumns(self,columns):
		'''
		columns are ismply list values. use self.sheet now. Note that all columns start from same row determined by the above
		'''

		
		
		for i in range(len(columns)):
			colid=i+self.startingColumnIndex
			col=self.sheet[self.columnLetters[colid]+str(self.columnsRow)]
			col.value=columns[i]
			col.font=self.colHeadersFont

	def __buildFileName(self):
		'''
		Just to have all exports standardized
		'''
		today=date.today()
		rands=''.join(random.choice(string.digits ) for _ in range(9))
		fileName='_'.join([self.calledfrom,self.current_user_name,self.append_to_file_name, str(today.year),str(today.month),str(today.day),str(rands)])
		
		fileNameFull=os.path.join(self.save_to,fileName)
	
		return [fileNameFull + '.xlsx','static/exports/xls/' + fileName + '.xlsx']
		




	def __inventory(self):
		'''
		Produce inventory records
		'''
		

		self.__workbook('Inventory')

		qColumns=['#','Product Name','Total Found','Store']

		dColumns=['#','Product Name','Tag','Store','Expires','Quantity','Original Price','Current Status']
		#build the columns here
		data=self.data
		if self.kind=='q':
			#the data is quantitied
			#build columns here first
			self.__buildColumns(qColumns)
			#now columns are built. So we start building records now

			row=self.startingRow
			column=self.startingColumnIndex

			record=1

			for d in range(len(self.data)):
				product_name=self.data[d]['item__product__name']
				manf=self.data[d]['item__manf__name']
				brand=self.data[d]['item__brand__name']
				store=self.data[d]['store__name']
				total_items=self.data[d]['total_items']
				measured_in=self.data[d]['item__product__measurement_unit']

				record_column=self.sheet[self.columnLetters[column] + str(row)]
				prod_column=self.sheet[self.columnLetters[column + 1] + str(row)]
				total_column=self.sheet[self.columnLetters[column + 2] +str(row)]
				store_column=self.sheet[self.columnLetters[column + 3] + str(row)]

				record_column.value=record
				prod_column.value=' '.join([product_name, manf,brand])
				total_column.value=' '.join([str(total_items),measured_in])
				store_column.value=store

				row=row + 1
				record=record +1
				#reset column
				column=self.startingColumnIndex


		elif self.kind=='d':
			#details.
			#build columns here first
			self.__buildColumns(dColumns)
			#now columns are built. So we start building records now

			row=self.startingRow
			column=self.startingColumnIndex

			record=1
			#items=ItemInStore.objects.filter(**filters).values().order_by('item__product__name',)
			#['#','Product Name','Tag','Store','Expires','Quantity','Original Price','Current Status']
			for d in range(len(self.data)):
				product_name=self.data[d]['item__product__name']
				manf=self.data[d]['item__manf__name']
				brand=self.data[d]['item__brand__name']
				store=self.data[d]['store__name']
				quantity=self.data[d]['quantity']
				measured_in=self.data[d]['item__product__measurement_unit']
				tag=self.data[d]['item__tag']
				price=self.data[d]['item__price']
				expire_on=self.data[d]['item__expire_on']
				status=self.data[d]['current_status']


				record_column=self.sheet[self.columnLetters[column] + str(row)]
				prod_column=self.sheet[self.columnLetters[column + 1] + str(row)]
				tag_column=self.sheet[self.columnLetters[column + 2] +str(row)]
				store_column=self.sheet[self.columnLetters[column + 3] + str(row)]
				expires_column=self.sheet[self.columnLetters[column + 4] + str(row)]
				quantity_column=self.sheet[self.columnLetters[column + 5] + str(row)]
				price_column=self.sheet[self.columnLetters[column + 6] + str(row)]
				status_column=self.sheet[self.columnLetters[column + 7] + str(row)]

				record_column.value=record
				prod_column.value=' '.join([product_name, manf,brand])
				quantity_column.value=' '.join([str(quantity),measured_in])
				store_column.value=store
				tag_column.value=tag 
				price_column.value=price
				expires_column.value=expire_on
				status_column.value=status

				row=row + 1
				record=record +1
				#reset column
				column=self.startingColumnIndex

		filename=self.__buildFileName()

		self.wb.save(filename[0])

		#return the file name now
		self.wb.close()

		return filename


	def __expiring(self):
		'''
		Produce expiring records
		'''
		'''
		self.data=data
		self.kind=datakind
		self.append_to_file_name=append_to_file_name
		self.extra=extra
		'''

		self.__workbook('Expiring')



		qColumns=['#','Product Name','Total Quantity','Store']

		dColumns=['#','Product Name','Store','Quantity','Original Price','Expire On']
		#build the columns here
		data=self.data

	
		if self.kind=='q':
			#the data is quantitied
			#build columns here first
			self.__buildColumns(qColumns)
			#now columns are built. So we start building records now

			row=self.startingRow
			column=self.startingColumnIndex

			record=1

			for d in range(len(self.data)):
				product_name=self.data[d]['item__product__name']
				manf=self.data[d]['item__manf__name']
				brand=self.data[d]['item__brand__name']
				store=self.data[d]['store__name']
				total_items=self.data[d]['total_items']
				measured_in=self.data[d]['item__product__measurement_unit']

				record_column=self.sheet[self.columnLetters[column] + str(row)]
				prod_column=self.sheet[self.columnLetters[column + 1] + str(row)]
				total_column=self.sheet[self.columnLetters[column + 2] +str(row)]
				store_column=self.sheet[self.columnLetters[column + 3] + str(row)]

				record_column.value=record
				prod_column.value=' '.join([product_name, manf,brand])
				total_column.value=' '.join([str(total_items),measured_in])
				store_column.value=store

				row=row + 1
				record=record +1
				#reset column
				column=self.startingColumnIndex


		elif self.kind=='d':
			#details.
			#build columns here first
			self.__buildColumns(dColumns)
			#now columns are built. So we start building records now

			row=self.startingRow
			column=self.startingColumnIndex

			record=1
			
			#dColumns=['#','Product Name','Store','Quantity','Original Price','Expire On']
			for d in range(len(self.data)):
				product_name=self.data[d]['item__product__name']
				manf=self.data[d]['item__manf__name']
				brand=self.data[d]['item__brand__name']
				store=self.data[d]['store__name']
				quantity=self.data[d]['quantity']
				measured_in=self.data[d]['item__product__measurement_unit']
				
				price=self.data[d]['item__price']
				expire_on=self.data[d]['item__expire_on']
				


				record_column=self.sheet[self.columnLetters[column] + str(row)]
				prod_column=self.sheet[self.columnLetters[column + 1] + str(row)]
				
				store_column=self.sheet[self.columnLetters[column + 2] + str(row)]
				expires_column=self.sheet[self.columnLetters[column + 5] + str(row)]
				quantity_column=self.sheet[self.columnLetters[column + 3] + str(row)]
				price_column=self.sheet[self.columnLetters[column + 4] + str(row)]
			

				record_column.value=record
				prod_column.value=' '.join([product_name, manf,brand])
				quantity_column.value=' '.join([str(quantity),measured_in])
				store_column.value=store
				
				price_column.value=price
				expires_column.value=expire_on
				

				row=row + 1
				record=record +1
				#reset column
				column=self.startingColumnIndex

		filename=self.__buildFileName()

		self.wb.save(filename[0])

		#return the file name now
		self.wb.close()

		return filename



	def __finishing_items(self):
		'''
		Produce finishing records
		'''
		

		self.__workbook('Finishing Items')

		qColumns=['#','Item','Kind','Total In Store','Min Value','Max Value','Required Amount']

		
		#build the columns here
		data=self.data
		
		#the data is quantitied
		#build columns here first
		self.__buildColumns(qColumns)
		#now columns are built. So we start building records now

		row=self.startingRow
		column=self.startingColumnIndex

		record=1

		for d in range(len(self.data)):

			product_name=self.data[d]['pro_name'] #Item
			pro_kind=self.data[d]['pro_kind'] #Kind
			total_items=self.data[d]['total_items'] #Total
			measured_in=self.data[d]['measure_unit']
			min_value=self.data[d]['min_value']
			max_value=self.data[d]['max_value']
			remaining=0

			

			remaining=min_value - total_items


			record_column=self.sheet[self.columnLetters[column] + str(row)]
			prod_column=self.sheet[self.columnLetters[column + 1] + str(row)]
			kind_column=self.sheet[self.columnLetters[column + 2] + str(row)]
			total_column=self.sheet[self.columnLetters[column + 3] +str(row)]
			min_value_column=self.sheet[self.columnLetters[column + 4] +str(row)]
			max_value_column=self.sheet[self.columnLetters[column + 5] +str(row)]
			remain_column=self.sheet[self.columnLetters[column +6] + str(row)]
			

			record_column.value=record
			prod_column.value=product_name
			kind_column.value=pro_kind

			total_column.value=' '.join([str(total_items),measured_in])
			min_value_column.value=' '.join([str(min_value),measured_in])
			max_value_column.value=' '.join([str(max_value),measured_in])
			remain_column.value=' '.join([str(remaining),measured_in])

			row=row + 1
			record=record +1
			#reset column
			column=self.startingColumnIndex


		

		filename=self.__buildFileName()

		self.wb.save(filename[0])

		#return the file name now
		self.wb.close()

		return filename
